import os
import openpyxl
import numpy as np
import math
import xlsxwriter as xw


##path = r"F:/learn/数模/jsjm/发布/B题"
path = r"D:/desktop/123/B题"
os.chdir(path)  # 修改工作路径

workbook1 = openpyxl.load_workbook('附件1-红方相关信息.xlsx')
workbook2 = openpyxl.load_workbook('附件2-蓝方相关信息.xlsx')	# 返回一个workbook数据类型的值
#print(workbook1.sheetnames)
#print(workbook2.sheetnames)	# 打印Excel表中的所有表

ServicesA = workbook1['A军种']
ServicesB = workbook1['B军种']
ServicesC = workbook1['C军种']

Target = workbook2['蓝方目标的相关数据']
Relation = workbook2['目标关联关系']


yuzhibiao=[
    [0]*15,
    [0]*15,
    [0]*15,
    [0]*15,
    [0]*25,
    [0]*25,
    [0]*25,
    [0]*25,
    [0]*25,
    [0]*25,
    [0]*10,
    [0]*10,
    [0]*8,
    [0]*8,
    [0]*8,
    [0]*8,
    [0]*3,
    [0]*3,
    [0]*3,
    [0]*3,
    ]

#剩余弹药表格
shengyudanyaobiao=[
    [8]*15,
    [8]*15,
    [12]*15,
    [12]*15,
    [270]*25,
    [270]*25,
    [270]*25,
    [270]*25,
    [270]*25,
    [270]*25,
    [1]*10,
    [1]*10,
    [6]*8,
    [0]*8,
    [0]*8,
    [0]*8,
    [0]*3,
    [0]*3,
    [0]*3,
    [0]*3,  
    ]




biao = [[] for i in range(20)]



r = np.zeros((148, 148))
cell = Relation['3:73']
for i in cell:
    x = i[0].value
    y = i[1].value
    r[x][y] = 1

tar_info = []
num = 0
cell = Target['2:1775']
a = {}
for i in cell:
    if i[0].value != num:
        if num != 0:
            tar_info.append(a)
            a = {}
        num = i[0].value
        a['num'] = num
        if i[1].value == '是':
            a['set'] = 1 
        else:
            a['set'] = 0
        a['local'] = i[2].value
        ammo = i[6].value
        a[ammo] = [i[4].value, float(i[5].value), i[9].value]
    else:
        ammo = i[6].value
        a[ammo] = [i[4].value, float(i[5].value), i[9].value]
for i in tar_info:
    ans = i['local'][1:-1].split(',')
    x = int(ans[0])
    y = int(ans[1])
    i['x'] = x
    i['y'] = y


cell = ServicesA['A4:H8']
platformA = []
for i in cell:
    a = {}
    a['type'] = i[0].value
    a['ammo'] = i[1].value
    a['missile_max'] = i[2].value
    a['target_max'] = i[3].value
    a['limit'] = i[4].value
    if i[4].value == '-':
        a['limit'] = 0
    a['price'] = i[5].value
    a['number'] = i[6].value
    a['yuzhi'] = i[7].value
    platformA.append(a)


cell = ServicesA['K4:N8']
ammoA = []
for i in cell:
    a = {}
    a['ammo'] = i[0].value
    a['number'] = i[1].value
    a['price'] = i[2].value
    a['range'] = int(i[3].value)
    ammoA.append(a)


cell = ServicesA['P4:S15']
regionA = []
for i in cell:
    a={}
    a['ID'] = i[0].value
    a['local'] = i[1].value
    a[i[2].value] = i[3].value
    regionA.append(a)



cell = ServicesB['A3:I7']
platformB = []
for i in cell:
    a = {}
    a['type'] = i[0].value
    a['ammo'] = i[1].value
    a['missile_max'] = i[2].value
    a['target_max'] = i[3].value
    a['limit'] = i[4].value
    if i[4].value == '-':
        a['limit'] = 0
    a['radius'] = i[5].value
    a['price'] = i[6].value
    a['number'] = i[7].value
    a['yuzhi'] = i[8].value
    platformB.append(a)
platformB[1]['number'] = 6
platformB[1]['yuzhi'] = 0.3
platformB[3]['number'] = 10
platformB[3]['yuzhi'] = 0.3
platformB[4]['number'] = 10
platformB[4]['yuzhi'] = 0.3



cell = ServicesB['L3:O7']
ammoB = []
for i in cell:
    a = {}
    a['ammo'] = i[0].value
    a['number'] = i[1].value
    a['price'] = i[2].value
    a['range'] = int(i[3].value)
    ammoB.append(a)


cell = ServicesB['Q3:T6']
regionB = []
for i in cell:
    a={}
    a['ID'] = i[0].value
    a['local'] = i[1].value
    a[i[2].value] = i[3].value
    regionB.append(a)
regionB[1]['local'] = '(9000,4700)'
regionB[3]['local'] = '(9500,3500)'




cell = ServicesC['A4:H5']
platformC = []
for i in cell:
    a = {}
    a['type'] = i[0].value
    a['ammo'] = i[1].value
    a['missile_max'] = i[2].value
    a['target_max'] = i[3].value
    a['limit'] = i[4].value
    if i[4].value == '-':
        a['limit'] = 0
    a['price'] = i[5].value
    a['number'] = i[6].value
    a['yuzhi'] = i[7].value
    platformC.append(a)
platformC[0]['ammo'] = ['HD1', 'HD2']


cell = ServicesC['K4:N5']
ammoC = []
for i in cell:
    a = {}
    a['ammo'] = i[0].value
    a['number'] = i[1].value
    a['price'] = i[2].value
    a['range'] = int(i[3].value)
    ammoC.append(a)


cell = ServicesC['P4:S7']
regionC = []
for i in cell:
    a={}
    a['ID'] = i[0].value
    a['local'] = i[1].value
    a[i[2].value] = i[3].value
    regionC.append(a)
regionC[1]['local'] = '(8600,2600)'
regionC[3]['local'] = '(8800,2800)'



ammo = ammoA + ammoB + ammoC
platform = platformA + platformB + platformC
region = regionA + regionB + regionC
for i in region:
    ans = i['local'][1:-1].split(',')
    x = int(ans[0])
    y = int(ans[1])
    i['x'] = x
    i['y'] = y



ammo_list = ['LD1', 'LD2', 'LD3', 'LD4', 'LD5', 'KD1', 'KD2', 'KD3', 'KD4', 'KD5', 'HD1', 'HD2']


tar = []
for i in tar_info:
    a = {}
    a['num'] = i['num']
    a['set'] = i['set']
    x = i['x']
    y = i['y']
    for key in i:#key是导弹类型
        if key in ammo_list:
            t = i[key][0]#t是平台类型
            for r in region:
                if t in r.keys():
                    dis = math.sqrt((r['x']-x)**2+(r['y']-y)**2)
                    ran = 0
                    radius = 0
                    for am in ammo:
                        if am['ammo'] == key:
                            ran = am['range']
                    for k in platform:
                        if t in k['type'] and key in k['ammo'] and 'radius' in k.keys():
                            radius = k['radius']
                    if dis <= ran+radius:
                        a[key] = [t, r['ID'], i[key][1], i[key][2]]
    tar.append(a)
                        
                            




platform_num = {'L1':26, 'L2':20, 'L3':6, 'L4':6, 'L5':16, 'K1':6, 'K2':10, 'H1':2, 'H2':2}
region_num = [15, 15, 15, 15, 25, 25, 25, 25, 25, 25, 10, 10, 8, 8, 8, 8, 3, 3, 3, 3]
ammo_num = {'LD1':1382, 'LD2':1728, 'LD3':6480, 'LD4':6480, 'LD5':44, 'KD1':81, 'KD2':54, 'KD3':90, 'KD4':90, 'KD5':180, 'HD1':16, 'HD2':48}
platform_price = {'L1':500, 'L2':300, 'L3':200, 'L4':150, 'L5':900, 'K1':600, 'K2':600, 'H1':1000, 'H2':800}
ammo_price = {'LD1':0.5, 'LD2':0.1, 'LD3':0.03, 'LD4':0.01, 'LD5':25, 'KD1':5, 'KD2':12, 'KD3':3, 'KD4':2.6, 'KD5':0.66, 'HD1':2, 'HD2':4}
most_missile = {'LD1':8, 'LD2':12, 'LD3':270, 'LD4':270, 'LD5':1, 'KD1':6, 'KD2':4, 'KD3':4, 'KD4':4, 'KD5':8, 'HD1':16, 'HD2':16}
most_yuzhi = {'L1':0.6, 'L2':0.6, 'L3':0.6, 'L4':0.6, 'L5':0.6, 'K1':0.3, 'K2':0.3, 'H1':0.7, 'H2':0.7}
A = ['LD1', 'LD2', 'LD3', 'LD4', 'LD5']
B = ['KD1', 'KD2', 'KD3', 'KD4', 'KD5']
C = ['HD1', 'HD2']


#问题一
tar1 = tar[0:50]
tar1.sort(key=lambda s: s['set'], reverse = True)
ans = []
time = 0
while len(tar1) != 0:
    m = np.zeros((20, 50))
    re = []
    time+= 1
    print(time)
    if time == 40:
        print(132)
    for i in tar1:
        i['set'] = 0
        ans0 = 'NULL'   ##最便宜导弹型号
        number0 = 0
        cost0 = 999999999
        for key in i:#key是导弹
            if key in ammo_list:
                
                number0 = math.ceil(i[key][3]/most_missile[key])##摧毁该目标需要的平台数量
                money = ammo_price[key] * i[key][3] + platform_price[i[key][0]] *i[key][2] *number0##该次摧毁成本
                if number0 <= platform_num[i[key][0]] and i[key][3] <= ammo_num[key] and number0 <= region_num[i[key][1]-1]:##平台总数约束 + 导弹总数约束 + 部署数量约束
                    m[i[key][1]-1][i['num']-1] = 1 ##可以打击
                if money <= cost0 and number0 <= platform_num[i[key][0]] and i[key][3] <= ammo_num[key] and number0 <= region_num[i[key][1]-1]:##更新最小成本
                    cost0 = money
                    ans0 = key
        # 打击目标序号 导弹型号 打击平台种类 打击平台区域 打击代价 所需平台数量 所需导弹数量 阈值
        re.append([i['num'], ans0, i[ans0][0], i[ans0][1], cost0, number0, i[ans0][3], i[ans0][2]])
        #platform_num[i[ans0][0]] -= number0
        #region_num[i[ans0][1]-1] -= number0
        #ammo_num[ans0] -= i[ans0][3]
        # if s == 1:
        #     ans1 = 'NULL'
        #     number1 = 0
        #     cost1 = 999999999
        #     for key in i:
        #         if((ans0 in A and key not in A) or (ans0 in B and key not in B) or (ans0 in C and key not in C)) and (key in ammo_list):
        #             number1 = math.ceil(i[key][3]/most_missile[key])
        #             money = ammo_price[key] * i[key][3] * number0 + platform_price[i[key][0]] *i[key][2]
        #             if money <= cost1 and number1 <= platform_num[i[key][0]] and i[key][3] <= ammo_num[key] and number1 <= region_num[i[key][1]-1]:
        #                 cost1 = money
        #                 ans1 = key
        #     re.append([i['num'], ans1, i[ans1][0], i[ans1][1], cost1, number1, i[ans0][3], i[ans0][2], 'spare'])
            #platform_num[i[ans1][0]] -= number1
            #region_num[i[ans1][1]-1] -= number1
            #ammo_num[ans1] -= i[ans1][3]
    
    
    warning = 0
    for i in range(50):
        num = 0
        for j in range(20):
            if m[j][i] == 1:
                num += 1
        if num == 1:
            warning = i+1
            break
    
    if warning != 0:
        choose = 0
        cost = re[0][4]
        for i in range(len(re)):
            if re[i][0] == warning:
                choose = i
                cost = re[i][4]
        platform_num[re[choose][2]] -= re[choose][5]
        region_num[re[choose][3]-1] -= re[choose][5]
        ammo_num[re[choose][1]] -= re[choose][6]
    
        ID = re[choose][0]
        for i in tar1:
            if i['num'] == ID:
                if i['set'] == 1:
                    i['set'] = 0
                else:
                    tar1.remove(i)
        
        
        ans.append(re[choose])
        continue
    
    
    
    
    
    
    # 打击目标序号 导弹型号 打击平台种类 打击平台区域 打击代价 所需平台数量 所需导弹数量  阈值      
    
    choose = 0
    cost = re[0][4]
    for i in range(1,len(re)):
        if re[i][4] < cost:
            choose = i
            cost = re[i][4]
    if len(biao[re[choose][3]-1]) == 0:
        # 阈值 剩余导弹数 导弹种类 平台种类
        if re[choose][7] < most_yuzhi[re[choose][2]]:
            biao[re[choose][3]-1] + [re[choose][7], re[choose][5]*most_missile[re[choose][1]], re[choose][1], re[choose][2]]
    else:
        if biao[re[choose][3]-1][1] > re[choose][6] and biao[re[choose][3]-1][0] + re[choose][7] < most_yuzhi[re[choose][2]]:
            biao[re[choose][3]-1][1] -= re[choose][6]
            biao[re[choose][3]-1][0] += re[choose][7]
            re[choose][5] = 0
            re[choose][6] = 0
        else:
            #一个平台最多用两次
            if re[choose][6] - (re[choose][5]-1) * most_missile[re[choose][1]] <= biao[re[choose][3]-1][1]:
                biao[re[choose][3]-1] = []
                re[choose][5] -= 1
                re[choose][6] -= 1
            else:
                biao[re[choose][3]-1] + [re[choose][7], most_missile[re[choose][1]] - (re[choose][6] - (re[choose][5]-1) * most_missile[re[choose][1]]), re[choose][1], re[choose][2]]
    platform_num[re[choose][2]] -= re[choose][5]
    region_num[re[choose][3]-1] -= re[choose][5]
    ammo_num[re[choose][1]] -= re[choose][6]
    
    
    ID = re[choose][0]
    for i in tar1:
        if i['num'] == ID:
            if i['set'] == 1:
                i['set'] = 0
            else:
                tar1.remove(i)
    
    
    ans.append(re[choose])
    
    






